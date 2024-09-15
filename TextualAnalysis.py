import os
import openpyxl
import pandas as pd
import nltk
import re


# Paths to necessary files and directories
input_folder = 'articles'
output_path = 'Output Data Structure.xlsx'
stop_words_dir = 'StopWords'
master_dict_path = 'MasterDictionary'

# Storing stop words
stop_words = set()
for filename in os.listdir(stop_words_dir):
    if filename.endswith('.txt'):
        file_path = os.path.join(stop_words_dir, filename)
        with open(file_path, 'r') as file:
            for line in file:
                word = line.strip()
                if word:
                    stop_words.add(word.lower())

# Loading  positive and negative words from the master dictionary
positive_words = set()
negative_words = set()
with open(os.path.join(master_dict_path, 'positive-words.txt')) as f:
    positive_words = set([line.strip() for line in f if line.strip() and line.strip() not in stop_words])
with open(os.path.join(master_dict_path, 'negative-words.txt')) as f:
    negative_words = set([line.strip() for line in f if line.strip() and line.strip() not in stop_words])


# function to calculate syllables in a word
def syllable_count(word):
    word = word.lower()
    vowels = "aeiou"
    count = sum(1 for char in word if char in vowels)
    if word.endswith("es") or word.endswith("ed"):
        count -= 1
    return max(1, count)


# function to analyze text and calculate required metrics
def analyze_text(text):
    # Tokenize text
    sentences = nltk.sent_tokenize(text)
    words = nltk.word_tokenize(text)
    cleaned_words = [word for word in words if word.lower() not in stop_words and word.isalnum()]

    # Calculate positive and negative scores
    positive_score = sum(1 for word in cleaned_words if word in positive_words)
    negative_score = sum(1 for word in cleaned_words if word in negative_words)

    polarity_score = (positive_score - negative_score) / ((positive_score + negative_score) + 0.000001)
    subjectivity_score = (positive_score + negative_score) / (len(cleaned_words) + 0.000001)

    # Readability metrics
    avg_sentence_length = len(cleaned_words) / len(sentences)
    complex_words = [word for word in cleaned_words if syllable_count(word) > 2]
    percentage_complex_words = len(complex_words) / len(cleaned_words)
    fog_index = 0.4 * (avg_sentence_length + percentage_complex_words)

    avg_words_per_sentence = len(cleaned_words) / len(sentences)
    complex_word_count = len(complex_words)
    word_count = len(cleaned_words)
    syllables_per_word = sum(syllable_count(word) for word in cleaned_words) / len(cleaned_words)

    # Personal pronouns count
    personal_pronouns = len(re.findall(r'\b(I|we|my|ours|us)\b', text, re.I))

    avg_word_length = sum(len(word) for word in cleaned_words) / len(cleaned_words)

    return {
        'positive_score': positive_score,
        'negative_score': negative_score,
        'polarity_score': polarity_score,
        'subjectivity_score': subjectivity_score,
        'avg_sentence_length': avg_sentence_length,
        'percentage_complex_words': percentage_complex_words,
        'fog_index': fog_index,
        'avg_words_per_sentence': avg_words_per_sentence,
        'complex_word_count': complex_word_count,
        'word_count': word_count,
        'syllables_per_word': syllables_per_word,
        'personal_pronouns': personal_pronouns,
        'avg_word_length': avg_word_length
    }


wb = openpyxl.load_workbook(output_path)
ws = wb.active
headers = [cell.value for cell in ws[1]]
header_count = len(headers)
# Iterate over each file in the input folder
for file_name in os.listdir(input_folder):
    if file_name.endswith('.txt'):
        file_path = os.path.join(input_folder, file_name)

        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        # calling analyze_text function
        analysis_results = analyze_text(content)
        url_id = file_name.split('.')[0]
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            if row[0].value == url_id:
                for col_idx, (key, value) in enumerate(analysis_results.items(), start=2):
                    if col_idx < header_count-2:
                        row[col_idx].value = value
                break
wb.save(output_path)
print("Article extraction, analysis, and saving completed.")
